import time
import sys
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
PACKAGE_DIR = BASE_DIR / ".packages"
if PACKAGE_DIR.exists():
    sys.path.insert(0, str(PACKAGE_DIR))

import pulp
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUTPUT_FILE = BASE_DIR / "results" / "pulp_same_case_results.xlsx"

SHIFT_LENGTH = 8 * 60
MAX_DAYS = 5
HORIZON = SHIFT_LENGTH * MAX_DAYS
BIG_M = HORIZON
MAINTENANCE_USAGE_LIMIT = 180
MAINTENANCE_CONDITION_DURATION = 30

SAME_CASE_ORDERS = [
    {"order_id": "O1", "batch_id": "O1_B1", "product_id": "P1", "quantity": 10},
    {"order_id": "O2", "batch_id": "O2_B1", "product_id": "P2", "quantity": 8},
    {"order_id": "O3", "batch_id": "O3_B1", "product_id": "P1", "quantity": 6},
    {"order_id": "O4", "batch_id": "O4_B1", "product_id": "P3", "quantity": 12},
]

MAINTENANCE_STRESS_ORDERS = [
    {"order_id": "S1", "batch_id": "S1_B1", "product_id": "P1", "quantity": 40},
    {"order_id": "S2", "batch_id": "S2_B1", "product_id": "P2", "quantity": 35},
    {"order_id": "S3", "batch_id": "S3_B1", "product_id": "P1", "quantity": 30},
    {"order_id": "S4", "batch_id": "S4_B1", "product_id": "P3", "quantity": 32},
]

PRODUCT_MACHINE_TIMES = {
    "P1": {"Cutting": 5, "Drilling": 3, "Painting": 4},
    "P2": {"Cutting": 4, "Milling": 6, "Painting": 5},
    "P3": {"Drilling": 4, "Welding": 7, "Painting": 3},
}

MACHINES = ["Cutting", "Drilling", "Milling", "Welding", "Painting"]

SAME_CASE_DUE_DATES = {
    "O1": 480,
    "O2": 480,
    "O3": 960,
    "O4": 960,
}

MAINTENANCE_STRESS_DUE_DATES = {
    "S1": 960,
    "S2": 960,
    "S3": 1440,
    "S4": 1440,
}

SCENARIOS = [
    {
        "name": "same_case",
        "sheet": "same_case",
        "orders": SAME_CASE_ORDERS,
        "due_dates": SAME_CASE_DUE_DATES,
    },
    {
        "name": "maintenance_stress_case",
        "sheet": "maint_stress",
        "orders": MAINTENANCE_STRESS_ORDERS,
        "due_dates": MAINTENANCE_STRESS_DUE_DATES,
    },
]

MAINTENANCE_WINDOWS = {
    "Cutting": [(120, 60)],
    "Painting": [(300, 60)],
}


def build_operations(orders):
    operations = []

    for order in orders:
        for machine, cycle_time in PRODUCT_MACHINE_TIMES[order["product_id"]].items():
            operations.append(
                {
                    "operation_id": len(operations),
                    "order_id": order["order_id"],
                    "batch_id": order["batch_id"],
                    "product_id": order["product_id"],
                    "machine": machine,
                    "quantity": order["quantity"],
                    "duration": cycle_time * order["quantity"],
                }
            )

    return operations


def add_no_overlap(model, start_vars, operations, name, counter):
    for left_index in range(len(operations)):
        for right_index in range(left_index + 1, len(operations)):
            left = operations[left_index]
            right = operations[right_index]
            i = left["operation_id"]
            j = right["operation_id"]

            i_before_j = pulp.LpVariable(f"{name}_{i}_before_{j}", cat="Binary")
            counter["binary_variables"] += 1

            model += start_vars[i] + left["duration"] <= start_vars[j] + BIG_M * (1 - i_before_j)
            model += start_vars[j] + right["duration"] <= start_vars[i] + BIG_M * i_before_j


def add_maintenance_constraints(model, start_vars, operations, counter):
    for op in operations:
        machine = op["machine"]

        for window_index, (window_start, window_duration) in enumerate(MAINTENANCE_WINDOWS.get(machine, [])):
            i = op["operation_id"]
            window_end = window_start + window_duration
            before_window = pulp.LpVariable(f"op_{i}_before_maintenance_{window_index}", cat="Binary")
            counter["binary_variables"] += 1

            model += start_vars[i] + op["duration"] <= window_start + BIG_M * (1 - before_window)
            model += start_vars[i] >= window_end - BIG_M * before_window


def calculate_bottleneck(schedule_rows, makespan):
    machine_loads = {}
    for row in schedule_rows:
        machine = row["machine"]
        machine_loads[machine] = machine_loads.get(machine, 0) + row["duration"]

    sorted_loads = sorted(machine_loads.items(), key=lambda item: (-item[1], item[0]))
    ranks = {}
    current_rank = 0
    previous_load = None

    for machine, load in sorted_loads:
        if load != previous_load:
            current_rank += 1
            previous_load = load
        ranks[machine] = current_rank

    return [
        {
            "machine": machine,
            "machine_load_min": load,
            "makespan_min": makespan,
            "utilization_percent": round(100 * load / makespan, 2),
            "bottleneck_rank": ranks[machine],
        }
        for machine, load in sorted_loads
    ]


def solve_scenario(scenario):
    orders = scenario["orders"]
    due_dates = scenario["due_dates"]
    operations = build_operations(orders)
    counter = {"binary_variables": 0}
    op_day = {}
    condition_usage_vars = {}
    condition_maintenance_vars = {}

    model = pulp.LpProblem(f"open_shop_pulp_{scenario['name']}", pulp.LpMinimize)

    start_vars = {
        op["operation_id"]: pulp.LpVariable(f"start_{op['operation_id']}", 0, HORIZON, cat="Integer")
        for op in operations
    }
    end_vars = {
        op["operation_id"]: pulp.LpVariable(f"end_{op['operation_id']}", 0, HORIZON, cat="Integer")
        for op in operations
    }
    makespan = pulp.LpVariable("makespan", 0, HORIZON, cat="Integer")

    for op in operations:
        i = op["operation_id"]
        model += end_vars[i] == start_vars[i] + op["duration"]
        model += makespan >= end_vars[i]

    for machine in MACHINES:
        machine_ops = [op for op in operations if op["machine"] == machine]
        add_no_overlap(model, start_vars, machine_ops, f"machine_{machine}", counter)

    for order in orders:
        order_ops = [op for op in operations if op["order_id"] == order["order_id"]]
        add_no_overlap(model, start_vars, order_ops, f"order_{order['order_id']}", counter)

    for op in operations:
        i = op["operation_id"]
        day_choices = []

        for day in range(MAX_DAYS):
            in_day = pulp.LpVariable(f"op_{i}_day_{day + 1}", cat="Binary")
            op_day[(i, day)] = in_day
            counter["binary_variables"] += 1
            day_choices.append(in_day)

            model += start_vars[i] >= day * SHIFT_LENGTH - BIG_M * (1 - in_day)
            model += end_vars[i] <= (day + 1) * SHIFT_LENGTH + BIG_M * (1 - in_day)

        model += sum(day_choices) == 1

    for machine in MACHINES:
        machine_ops = [op for op in operations if op["machine"] == machine]

        for day in range(MAX_DAYS):
            daily_usage = pulp.LpVariable(
                f"usage_{machine}_day_{day + 1}",
                lowBound=0,
                upBound=SHIFT_LENGTH,
                cat="Integer",
            )
            maintenance_required = pulp.LpVariable(
                f"condition_maintenance_{machine}_day_{day + 1}",
                cat="Binary",
            )
            counter["binary_variables"] += 1
            condition_usage_vars[(machine, day)] = daily_usage
            condition_maintenance_vars[(machine, day)] = maintenance_required

            model += daily_usage == sum(
                op["duration"] * op_day[(op["operation_id"], day)]
                for op in machine_ops
            )
            model += daily_usage <= MAINTENANCE_USAGE_LIMIT + BIG_M * maintenance_required
            model += daily_usage >= MAINTENANCE_USAGE_LIMIT + 1 - BIG_M * (1 - maintenance_required)
            model += daily_usage + MAINTENANCE_CONDITION_DURATION * maintenance_required <= SHIFT_LENGTH

    add_maintenance_constraints(model, start_vars, operations, counter)

    tardiness_vars = {}
    for order in orders:
        order_id = order["order_id"]
        completion = pulp.LpVariable(f"completion_{order_id}", 0, HORIZON, cat="Integer")
        tardiness = pulp.LpVariable(f"tardiness_{order_id}", 0, HORIZON, cat="Integer")
        tardiness_vars[order_id] = tardiness

        for op in operations:
            if op["order_id"] == order_id:
                model += completion >= end_vars[op["operation_id"]]

        model += tardiness >= completion - due_dates[order_id]

    model += makespan * 1000 + sum(tardiness_vars.values())

    solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=15)

    start_time = time.perf_counter()
    status = model.solve(solver)
    solve_time = time.perf_counter() - start_time

    schedule = []
    for op in operations:
        i = op["operation_id"]
        start = int(round(pulp.value(start_vars[i])))
        end = int(round(pulp.value(end_vars[i])))

        schedule.append(
            {
                "scenario": scenario["name"],
                **op,
                "start_min": start,
                "end_min": end,
                "start_day": start // SHIFT_LENGTH + 1,
                "end_day": (end - 1) // SHIFT_LENGTH + 1,
            }
        )

    schedule = sorted(schedule, key=lambda row: (row["start_min"], row["machine"]))
    final_makespan = int(round(pulp.value(makespan)))
    bottleneck = calculate_bottleneck(schedule, final_makespan)
    for row in bottleneck:
        row["scenario"] = scenario["name"]

    condition_maintenance = []
    for machine in MACHINES:
        for day in range(MAX_DAYS):
            key = (machine, day)
            required = int(round(pulp.value(condition_maintenance_vars[key])))
            condition_maintenance.append(
                {
                    "scenario": scenario["name"],
                    "solver": "PuLP CBC",
                    "machine": machine,
                    "day": day + 1,
                    "daily_usage_min": int(round(pulp.value(condition_usage_vars[key]))),
                    "threshold_min": MAINTENANCE_USAGE_LIMIT,
                    "maintenance_required": required,
                    "maintenance_reserved_min": MAINTENANCE_CONDITION_DURATION if required == 1 else 0,
                }
            )

    maintenance_triggers = sum(row["maintenance_required"] for row in condition_maintenance)
    reserved_maintenance = sum(row["maintenance_reserved_min"] for row in condition_maintenance)
    max_daily_usage = max(row["daily_usage_min"] for row in condition_maintenance)

    summary = [
        {
            "scenario": scenario["name"],
            "solver": "PuLP CBC",
            "status": pulp.LpStatus[status],
            "orders": len(orders),
            "operations": len(operations),
            "makespan_min": final_makespan,
            "total_tardiness_min": int(round(sum(pulp.value(v) for v in tardiness_vars.values()))),
            "maintenance_triggers": maintenance_triggers,
            "reserved_maintenance_min": reserved_maintenance,
            "max_daily_usage_min": max_daily_usage,
            "binary_variables": counter["binary_variables"],
            "solve_time_sec": round(solve_time, 3),
            "formulation": "MILP with big-M and binary ordering variables",
        }
    ]

    return summary, schedule, bottleneck, condition_maintenance


def write_sheet(workbook, sheet_name, rows):
    worksheet = workbook.create_sheet(sheet_name)
    if not rows:
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    for column_index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows]
        width = min(max(len(value) for value in values) + 2, 35)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def save_workbook(output_file, summary_rows, scenario_results):
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(workbook, "Summary", summary_rows)

    for scenario, _, schedule, bottleneck, condition_maintenance in scenario_results:
        suffix = scenario["sheet"]
        write_sheet(workbook, f"Schedule_{suffix}", schedule)
        write_sheet(workbook, f"Bottleneck_{suffix}", bottleneck)
        write_sheet(workbook, f"Condition_{suffix}", condition_maintenance)

    workbook.save(output_file)


def print_summary(summary_rows):
    headers = list(summary_rows[0].keys())
    print("\t".join(headers))
    for row in summary_rows:
        print("\t".join(str(row.get(header, "")) for header in headers))


def main():
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    results = []
    for scenario in SCENARIOS:
        print(f"Solving {scenario['name']} with PuLP...")
        results.append((scenario, *solve_scenario(scenario)))

    summary = []
    for _, scenario_summary, _, _, _ in results:
        summary.extend(scenario_summary)

    save_workbook(OUTPUT_FILE, summary, results)

    print_summary(summary)
    print(f"\nExcel file created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
